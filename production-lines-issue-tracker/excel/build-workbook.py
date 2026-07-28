"""Builds the Production Lines issue tracker workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment

OUT = '/home/user/test/production-lines-issue-tracker/excel/Production-Lines-Issue-Tracker.xlsx'
LAST = 200                      # rows of the ticket table that are pre-formatted
R = f'$2:${LAST}'

INK, INK2, MUTED = '12181D', '47535E', '7A8894'
ACCENT, RULE, BAND = '0E6E7D', 'D5DCE2', 'F7F9FA'
SEV = {                          # fill, font
    'S1 - Line Down': ('FAE7E4', 'C23B2E'),
    'S2 - Major':     ('F7EEDC', 'B57312'),
    'S3 - Minor':     ('E6EEF6', '3D6E9E'),
    'S4 - Cosmetic':  ('EDF0F2', '7A8894'),
}
OK_FILL, OK_FONT = 'E2F0E9', '2E7D5B'

A = 'Arial'
h1   = Font(name=A, size=15, bold=True, color=INK)
h2   = Font(name=A, size=11, bold=True, color=INK)
body = Font(name=A, size=10, color=INK)
dim  = Font(name=A, size=9,  color=MUTED)
hdr  = Font(name=A, size=9,  bold=True, color='FFFFFF')
big  = Font(name=A, size=22, bold=True, color=INK)
hdr_fill = PatternFill('solid', fgColor=ACCENT)
thin = Side(style='thin', color=RULE)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ------------------------------------------------------------------ Choices --
ch = wb.active
ch.title = 'Choices'
COLS = {
    'A': ('Ticket Type', ['Issue', 'Request']),
    'B': ('Category', ['Mechanical', 'Electrical', 'Controls / PLC', 'Quality defect',
                       'Safety', 'Material supply', 'Utilities', 'Changeover',
                       'Housekeeping', 'Other']),
    'C': ('Severity', ['S1 - Line Down', 'S2 - Major', 'S3 - Minor', 'S4 - Cosmetic']),
    'D': ('Status', ['New', 'Triaged', 'In Progress', 'Waiting on Parts', 'Resolved', 'Closed']),
    'E': ('Shift', ['A - Days', 'B - Afters', 'C - Nights']),
    'F': ('Line State', ['Running', 'Degraded', 'Down', 'Planned Downtime']),
}
for col, (title, vals) in COLS.items():
    c = ch[f'{col}1']; c.value = title; c.font = hdr; c.fill = hdr_fill
    for i, v in enumerate(vals, start=2):
        cell = ch[f'{col}{i}']; cell.value = v; cell.font = body
    ch.column_dimensions[col].width = 18
ch['H1'] = 'Edit these lists to change what the dropdowns offer.'
ch['H1'].font = Font(name=A, size=10, bold=True, color=ACCENT)
ch['H2'] = ('Add a value directly under the last one in a column and the dropdown on the '
            'Tickets sheet picks it up. Renaming a value does NOT update tickets already '
            'using the old wording, and the colour rules on the Tickets sheet match the '
            'Severity and Status wording exactly - change those under Home > Conditional '
            'Formatting if you reword them.')
ch['H2'].font = dim
ch['H2'].alignment = Alignment(wrap_text=True, vertical='top')
ch.merge_cells('H2:N8')
ch.column_dimensions['H'].width = 14

names = {'TicketTypes': 'Choices!$A$2:$A$3',  'Categories': 'Choices!$B$2:$B$11',
         'Severities':  'Choices!$C$2:$C$5',  'Statuses':   'Choices!$D$2:$D$7',
         'Shifts':      'Choices!$E$2:$E$4',  'LineStates': 'Choices!$F$2:$F$5'}
for n, rng in names.items():
    wb.defined_names.add(DefinedName(n, attr_text=rng))

# ---------------------------------------------------------------- Read me ----
rm = wb.create_sheet('Read me')
rm['A1'] = 'Production Lines 1-12 - Issue & Request Tracker'; rm['A1'].font = h1
rm['A2'] = 'Save this file to SharePoint or a Teams channel and everyone can edit it at once in the browser.'
rm['A2'].font = Font(name=A, size=10, color=INK2)

lines_txt = [
    ('How to use it', True),
    ('1.  Raise a ticket on the Tickets sheet - one row per issue or request. Fill the white cells; the grey ones calculate themselves.', False),
    ('2.  Pick from the dropdowns for Ticket Type, Line, Category, Severity, Status and Shift so the counts on the Dashboard stay accurate.', False),
    ('3.  Update Status as work progresses. Set Date Closed when it is finished - that is what drives "closed this week".', False),
    ('4.  The Dashboard sheet recalculates on its own. Nothing to press.', False),
    ('', False),
    ('Which cells you edit', True),
    ('White background = you type here.', False),
    ('Grey background = formula, leave it alone (Ticket ID and Age (days) on Tickets, everything on the Dashboard).', False),
    ('Row 2 of the Tickets sheet is an example. Delete it before you start.', False),
    ('', False),
    ('Severity - what the levels mean', True),
    ('S1 - Line Down    the line is stopped. Raise it here AND call the shift lead; do not wait for the ticket.', False),
    ('S2 - Major        running but degraded, quality or rate at risk. Same shift.', False),
    ('S3 - Minor        line running normally. Plan it into maintenance.', False),
    ('S4 - Cosmetic     improvement or backlog. Review at the weekly meeting.', False),
    ('', False),
    ('Changing the setup', True),
    ('Rename a line, change its owner or its state          -> Lines sheet', False),
    ('Add a category, status or severity                    -> Choices sheet', False),
    ('More than 200 tickets                                 -> copy the last row of the Tickets sheet downwards; the formulas and dropdowns come with it', False),
    ('', False),
    ('Worth knowing', True),
    ('Everyone who can open this file can edit any row, including other people\'s. A SharePoint list can restrict that per person; a spreadsheet cannot.', False),
    ('There are no automatic notifications. If you want an alert when an S1 is raised, that needs the SharePoint list version plus a Power Automate flow.', False),
    ('Attachments: paste a photo straight into the Details cell, or keep photos in a folder beside this file.', False),
]
r = 4
for text, is_head in lines_txt:
    c = rm.cell(row=r, column=1, value=text)
    c.font = h2 if is_head else Font(name=A, size=10, color=INK if not is_head else INK)
    if not is_head and text:
        c.font = Font(name=A, size=10, color=INK2)
    r += 1
rm.column_dimensions['A'].width = 132
rm.sheet_view.showGridLines = False

# ------------------------------------------------------------------ Lines ----
ln = wb.create_sheet('Lines')
ln['A1'] = 'Production lines'; ln['A1'].font = h1
ln['A2'] = 'Rename a line here and every dropdown and count follows. Line Code is what tickets store - keep the padding (L01, not L1) so it sorts correctly.'
ln['A2'].font = dim
ln.merge_cells('A2:E2')

heads = ['Line Code', 'Line Name', 'Line Owner', 'Line State', 'Open tickets']
for i, h in enumerate(heads, start=1):
    c = ln.cell(row=4, column=i, value=h); c.font = hdr; c.fill = hdr_fill; c.border = box
for i in range(1, 13):
    row = 4 + i
    ln.cell(row=row, column=1, value=f'L{i:02d}').font = Font(name=A, size=10, bold=True, color=INK)
    ln.cell(row=row, column=2, value=f'Line {i}').font = body
    ln.cell(row=row, column=3, value='').font = body
    ln.cell(row=row, column=4, value='Running').font = body
    f = (f'=SUMPRODUCT((Tickets!$D{R}=$A{row})*(Tickets!$G{R}<>"Resolved")'
         f'*(Tickets!$G{R}<>"Closed"))')
    c = ln.cell(row=row, column=5, value=f)
    c.font = body; c.fill = PatternFill('solid', fgColor=BAND)
    c.alignment = Alignment(horizontal='center')
    for col in range(1, 6):
        ln.cell(row=row, column=col).border = box
for col, w in zip('ABCDE', (12, 26, 24, 18, 14)):
    ln.column_dimensions[col].width = w
dv_state = DataValidation(type='list', formula1='=LineStates', allow_blank=True)
ln.add_data_validation(dv_state); dv_state.add(f'D5:D16')
ln.freeze_panes = 'A5'
wb.defined_names.add(DefinedName('LineCodes', attr_text='Lines!$A$5:$A$16'))

# ---------------------------------------------------------------- Tickets ----
tk = wb.create_sheet('Tickets')
cols = [
    ('Ticket ID', 10, True), ('Date Raised', 13, False), ('Ticket Type', 13, False),
    ('Line', 9, False), ('Category', 17, False), ('Severity', 16, False),
    ('Status', 16, False), ('Summary', 44, False), ('Details', 46, False),
    ('Reported By', 16, False), ('Shift', 12, False), ('Downtime (min)', 14, False),
    ('Assigned To', 16, False), ('Target Date', 13, False), ('Date Closed', 13, False),
    ('Resolution', 34, False), ('Age (days)', 11, True),
]
for i, (h, w, _calc) in enumerate(cols, start=1):
    c = tk.cell(row=1, column=i, value=h)
    c.font = hdr; c.fill = hdr_fill; c.border = box
    c.alignment = Alignment(horizontal='left', vertical='center')
    tk.column_dimensions[get_column_letter(i)].width = w
tk.row_dimensions[1].height = 22

tk['A1'].comment = Comment('Calculated - do not type here. Numbering starts at 1001.', 'Tracker')
tk['Q1'].comment = Comment('Calculated - days since Date Raised. Goes blank once the ticket is '
                           'Resolved or Closed.', 'Tracker')

example = ['', '=TODAY()-2', 'Issue', 'L03', 'Electrical', 'S1 - Line Down', 'In Progress',
           'Drive motor tripping under load',
           'Overload trips within ten minutes of a full-rate start. Line stopped.',
           'J. Okafor', 'B - Afters', 95, '', '', '', '', '']
for i, v in enumerate(example, start=1):
    if v != '':
        tk.cell(row=2, column=i, value=v)

calc_fill = PatternFill('solid', fgColor=BAND)
for row in range(2, LAST + 1):
    tk.cell(row=row, column=1,
            value=f'=IF($D{row}="","",1000+ROW()-1)').fill = calc_fill
    age = (f'=IF(OR($D{row}="",$B{row}=""),"",'
           f'IF(OR($G{row}="Resolved",$G{row}="Closed"),"",TODAY()-$B{row}))')
    tk.cell(row=row, column=17, value=age)
    tk.cell(row=row, column=17).fill = calc_fill
    for col in range(1, 18):
        c = tk.cell(row=row, column=col)
        c.font = body
        c.border = box
        c.alignment = Alignment(vertical='top', wrap_text=col in (8, 9, 16))
    for col in (2, 14, 15):
        tk.cell(row=row, column=col).number_format = 'yyyy-mm-dd'
    tk.cell(row=row, column=12).number_format = '#,##0'
    tk.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
    tk.cell(row=row, column=17).alignment = Alignment(horizontal='center', vertical='top')

for name, col in (('TicketTypes', 'C'), ('LineCodes', 'D'), ('Categories', 'E'),
                  ('Severities', 'F'), ('Statuses', 'G'), ('Shifts', 'K')):
    dv = DataValidation(type='list', formula1=f'={name}', allow_blank=True,
                        showDropDown=False, errorTitle='Pick from the list',
                        error='Use the dropdown so the Dashboard counts stay right. '
                              'Add new options on the Choices sheet.')
    tk.add_data_validation(dv)
    dv.add(f'{col}2:{col}{LAST}')

for label, (fill, font) in SEV.items():
    tk.conditional_formatting.add(
        f'F2:F{LAST}',
        FormulaRule(formula=[f'$F2="{label}"'], stopIfTrue=False,
                    fill=PatternFill('solid', fgColor=fill),
                    font=Font(name=A, size=10, bold=True, color=font)))
for label in ('Resolved', 'Closed'):
    tk.conditional_formatting.add(
        f'G2:G{LAST}',
        FormulaRule(formula=[f'$G2="{label}"'], stopIfTrue=False,
                    fill=PatternFill('solid', fgColor=OK_FILL),
                    font=Font(name=A, size=10, bold=True, color=OK_FONT)))
tk.conditional_formatting.add(
    f'G2:G{LAST}',
    FormulaRule(formula=['$G2="Waiting on Parts"'], stopIfTrue=False,
                fill=PatternFill('solid', fgColor=SEV['S2 - Major'][0]),
                font=Font(name=A, size=10, bold=True, color=SEV['S2 - Major'][1])))
# overdue: S1 older than a day, S2 older than three, anything older than a fortnight
tk.conditional_formatting.add(
    f'Q2:Q{LAST}',
    FormulaRule(formula=[f'AND($Q2<>"",OR(AND($F2="S1 - Line Down",$Q2>1),'
                         f'AND($F2="S2 - Major",$Q2>3),$Q2>14))'],
                stopIfTrue=False,
                fill=PatternFill('solid', fgColor=SEV['S1 - Line Down'][0]),
                font=Font(name=A, size=10, bold=True, color=SEV['S1 - Line Down'][1])))

tk.freeze_panes = 'B2'
tk.auto_filter.ref = f'A1:Q{LAST}'

# -------------------------------------------------------------- Dashboard ----
db = wb.create_sheet('Dashboard')
db['A1'] = 'Production Lines 1-12'; db['A1'].font = h1
db['A2'] = 'Every figure here is a formula over the Tickets sheet. Nothing to refresh.'
db['A2'].font = dim

OPEN = f'(Tickets!$G{R}<>"Resolved")*(Tickets!$G{R}<>"Closed")*(Tickets!$D{R}<>"")'
kpis = [
    ('Open tickets', f'=SUMPRODUCT({OPEN})', 'New, triaged, in progress, waiting'),
    ('Lines down (S1)', f'=SUMPRODUCT({OPEN}*(Tickets!$F{R}="S1 - Line Down"))', 'Stopped right now'),
    ('Oldest open (days)', f'=IFERROR(MAX(Tickets!$Q{R}),0)', 'Longest-running open ticket'),
    ('Closed last 7 days', f'=SUMPRODUCT((Tickets!$O{R}<>"")*(Tickets!$O{R}>=TODAY()-7))', 'By Date Closed'),
]
for i, (label, formula, note) in enumerate(kpis):
    col = 1 + i * 2
    c = db.cell(row=4, column=col, value=label.upper())
    c.font = Font(name=A, size=8, bold=True, color=MUTED)
    v = db.cell(row=5, column=col, value=formula); v.font = big
    n = db.cell(row=6, column=col, value=note); n.font = dim
    db.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    db.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)

db['A8'] = 'Open tickets by line'; db['A8'].font = h2
for i, h in enumerate(['Line', 'Name', 'Open', 'S1', 'S2', 'S3', 'S4'], start=1):
    c = db.cell(row=9, column=i, value=h); c.font = hdr; c.fill = hdr_fill; c.border = box
for i in range(1, 13):
    row, lrow = 9 + i, 4 + i
    db.cell(row=row, column=1, value=f'=Lines!$A{lrow}').font = Font(name=A, size=10, bold=True, color=INK)
    db.cell(row=row, column=2, value=f'=Lines!$B{lrow}').font = body
    per = f'(Tickets!$D{R}=$A{row})'
    db.cell(row=row, column=3, value=f'=SUMPRODUCT({OPEN}*{per})').font = body
    for j, sev in enumerate(SEV, start=4):
        db.cell(row=row, column=j,
                value=f'=SUMPRODUCT({OPEN}*{per}*(Tickets!$F{R}="{sev}"))').font = body
    for col in range(1, 8):
        c = db.cell(row=row, column=col); c.border = box
        if col >= 3:
            c.alignment = Alignment(horizontal='center')

db['I8'] = 'By status'; db['I8'].font = h2
for i, h in enumerate(['Status', 'Tickets'], start=9):
    c = db.cell(row=9, column=i, value=h); c.font = hdr; c.fill = hdr_fill; c.border = box
for i, st in enumerate(COLS['D'][1], start=10):
    db.cell(row=i, column=9, value=st).font = body
    db.cell(row=i, column=10,
            value=f'=SUMPRODUCT((Tickets!$G{R}="{st}")*(Tickets!$D{R}<>""))').font = body
    db.cell(row=i, column=10).alignment = Alignment(horizontal='center')
    db.cell(row=i, column=9).border = box; db.cell(row=i, column=10).border = box

db['I18'] = 'Open by category'; db['I18'].font = h2
for i, h in enumerate(['Category', 'Open'], start=9):
    c = db.cell(row=19, column=i, value=h); c.font = hdr; c.fill = hdr_fill; c.border = box
for i, cat in enumerate(COLS['B'][1], start=20):
    db.cell(row=i, column=9, value=cat).font = body
    db.cell(row=i, column=10,
            value=f'=SUMPRODUCT({OPEN}*(Tickets!$E{R}="{cat}"))').font = body
    db.cell(row=i, column=10).alignment = Alignment(horizontal='center')
    db.cell(row=i, column=9).border = box; db.cell(row=i, column=10).border = box

for col, w in zip('ABCDEFGHIJ', (10, 22, 9, 7, 7, 7, 7, 4, 20, 10)):
    db.column_dimensions[col].width = w
db.sheet_view.showGridLines = False
db['A31'] = ('Counts cover rows 2-200 of the Tickets sheet. If you extend past row 200, widen '
             'the ranges in these formulas to match.')
db['A31'].font = dim

order = ['Read me', 'Dashboard', 'Tickets', 'Lines', 'Choices']
wb._sheets = [wb[n] for n in order]
wb.active = 0
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('saved', OUT)
print('sheets:', wb.sheetnames)
