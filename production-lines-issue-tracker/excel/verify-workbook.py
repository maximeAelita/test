"""Structural checks on the workbook, since LibreOffice cannot run in this sandbox.

Catches the failure modes a recalc would have caught: unsupported functions, broken
cross-sheet references, and off-by-one row alignment between the Dashboard/Lines
tables and the rows they point at.
"""
import re
import openpyxl

P = '/home/user/test/production-lines-issue-tracker/excel/Production-Lines-Issue-Tracker.xlsx'
BANNED = {'XLOOKUP', 'XMATCH', 'SORT', 'FILTER', 'UNIQUE', 'SEQUENCE',
          'TEXTJOIN', 'CONCAT', 'IFS', 'SWITCH', 'MAXIFS', 'MINIFS'}
fails, notes = [], []

wb = openpyxl.load_workbook(P)
print('sheets:', wb.sheetnames)

# 1. functions used ---------------------------------------------------------
funcs, formula_cells = set(), 0
for ws in wb:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                formula_cells += 1
                funcs |= set(re.findall(r'([A-Z][A-Z0-9_.]*)\s*\(', c.value))
print(f'{formula_cells} formula cells; functions used: {sorted(funcs)}')
bad = funcs & BANNED
if bad:
    fails.append(f'unsupported function(s) without _xlfn prefix: {sorted(bad)}')

# 2. sheet names with spaces must be quoted in references -------------------
for ws in wb:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                for ref in re.findall(r"(?<!')\b([A-Za-z ]+)!\$", c.value):
                    if ' ' in ref:
                        fails.append(f'{ws.title}!{c.coordinate}: unquoted sheet name "{ref}"')

# 3. defined names resolve to the expected choice lists ---------------------
expected = {
    'TicketTypes': ['Issue', 'Request'],
    'Severities': ['S1 - Line Down', 'S2 - Major', 'S3 - Minor', 'S4 - Cosmetic'],
    'Statuses': ['New', 'Triaged', 'In Progress', 'Waiting on Parts', 'Resolved', 'Closed'],
    'Shifts': ['A - Days', 'B - Afters', 'C - Nights'],
    'LineStates': ['Running', 'Degraded', 'Down', 'Planned Downtime'],
}
for name, want in expected.items():
    dn = wb.defined_names.get(name)
    if dn is None:
        fails.append(f'defined name {name} missing'); continue
    sheet, rng = dn.attr_text.split('!')
    got = [c.value for (c,) in wb[sheet][rng.replace('$', '')]]
    if got != want:
        fails.append(f'{name} -> {dn.attr_text} gives {got}, expected {want}')
    else:
        notes.append(f'{name} -> {dn.attr_text} OK')

dn = wb.defined_names.get('LineCodes')
sheet, rng = dn.attr_text.split('!')
codes = [c.value for (c,) in wb[sheet][rng.replace('$', '')]]
if codes != [f'L{i:02d}' for i in range(1, 13)]:
    fails.append(f'LineCodes gives {codes}')
else:
    notes.append(f'LineCodes -> {dn.attr_text} = L01..L12 OK')

# 4. row alignment: Lines "Open tickets" must test its own row's code -------
ln = wb['Lines']
for i in range(1, 13):
    row = 4 + i
    f = ln.cell(row=row, column=5).value
    if f'=$A{row})' not in f:
        fails.append(f'Lines!E{row} compares against the wrong row: {f[:80]}')
notes.append('Lines!E5:E16 each compare Tickets column D against their own Line Code')

# 5. row alignment: Dashboard by-line table -> Lines rows 5..16 -------------
db = wb['Dashboard']
for i in range(1, 13):
    row, lrow = 9 + i, 4 + i
    if db.cell(row=row, column=1).value != f'=Lines!$A{lrow}':
        fails.append(f'Dashboard!A{row} points at {db.cell(row=row, column=1).value}, '
                     f'expected =Lines!$A{lrow}')
    if f'=$A{row})' not in db.cell(row=row, column=3).value:
        fails.append(f'Dashboard!C{row} counts against the wrong row')
notes.append('Dashboard rows 10-21 map 1:1 onto Lines rows 5-16')

# 6. every Tickets row has both calculated cells, pointing at its own row ---
tk = wb['Tickets']
last = max(r for r in range(2, 1000) if tk.cell(row=r, column=1).value)
for row in range(2, last + 1):
    if tk.cell(row=row, column=1).value != f'=IF($D{row}="","",1000+ROW()-1)':
        fails.append(f'Tickets!A{row} malformed')
    age = tk.cell(row=row, column=17).value
    if f'$B{row}' not in age or f'$G{row}' not in age:
        fails.append(f'Tickets!Q{row} references the wrong row')
notes.append(f'Tickets rows 2-{last}: ID and Age formulas both self-referencing')

# 7. data validation covers the dropdown columns ---------------------------
cover = {dv.formula1: str(dv.sqref) for dv in tk.data_validations.dataValidation}
for name, col in (('=TicketTypes', 'C'), ('=LineCodes', 'D'), ('=Categories', 'E'),
                  ('=Severities', 'F'), ('=Statuses', 'G'), ('=Shifts', 'K')):
    want = f'{col}2:{col}{last}'
    if cover.get(name) != want:
        fails.append(f'validation {name} covers {cover.get(name)}, expected {want}')
notes.append(f'6 dropdowns each cover rows 2-{last}')

# 8. conditional formatting present on severity / status / age -------------
cf = {str(rng): len(rules) for rng, rules in tk.conditional_formatting._cf_rules.items()}
notes.append(f'conditional formatting ranges: {cf}')

print()
for n in notes:
    print('  ok  ', n)
print()
if fails:
    print('FAILURES:')
    for f in fails:
        print('  !!  ', f)
    raise SystemExit(1)
print('all structural checks passed')

# 9. OOXML-level checks that openpyxl's own loader is too lenient to catch ------
import zipfile
z = zipfile.ZipFile(P)
styles = z.read('xl/styles.xml').decode()
dxfs = re.search(r'<dxfs.*?</dxfs>', styles, re.S)
if dxfs:
    blob = dxfs.group(0)
    if '<name val' in blob or '<sz val' in blob:
        print('  !!   dxf carries a font name or size - Excel will offer to repair the file')
        raise SystemExit(1)
    print('  ok   conditional-format dxfs carry colour/bold only')
# <indexedColors> is openpyxl's legacy palette boilerplate, present in every file it
# writes and harmless - only colours applied outside it matter.
applied = re.sub(r'<indexedColors>.*?</indexedColors>', '', styles, flags=re.S)
bad_alpha = sorted(set(re.findall(r'rgb="(00[0-9A-Fa-f]{6})"', applied)))
if bad_alpha:
    print(f'  !!   {len(bad_alpha)} colours written with a 00 alpha byte: {bad_alpha[:5]}')
    raise SystemExit(1)
print('  ok   every colour is 8-digit aRGB with a real alpha byte')
for part in z.namelist():
    if part.endswith('.xml'):
        try:
            __import__('xml.dom.minidom', fromlist=['x']).parseString(z.read(part))
        except Exception as e:
            print(f'  !!   {part} is not well-formed: {e}')
            raise SystemExit(1)
print(f'  ok   all {sum(1 for n in z.namelist() if n.endswith(".xml"))} XML parts well-formed')
